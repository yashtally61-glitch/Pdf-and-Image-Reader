import streamlit as st
import pandas as pd
import base64
import json
import re
import io
import os
import tempfile
import gc
import time
import requests
from PIL import Image, ImageEnhance
try:
    from pdf2image import convert_from_bytes, convert_from_path
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
from difflib import get_close_matches

st.set_page_config(page_title="DataScan → Excel", page_icon="⚡", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;700&display=swap');

    /* ── Base ── */
    .stApp { background-color: #f0f4f8; }
    .block-container { padding-top: 1.5rem !important; }

    /* ── Sidebar ── */
    section[data-testid="stSidebar"] { background: #1e2a3a !important; }
    section[data-testid="stSidebar"] * { color: #d0dce8 !important; }
    section[data-testid="stSidebar"] h3 { color: #60c8f0 !important; font-family: 'JetBrains Mono', monospace !important; font-size:0.85rem !important; }
    section[data-testid="stSidebar"] .stTextInput input,
    section[data-testid="stSidebar"] .stTextArea textarea {
        background: #243040 !important; border: 1px solid #3a4f60 !important;
        color: #e8f0f8 !important; font-family: 'JetBrains Mono', monospace !important; font-size:0.8rem !important;
    }
    section[data-testid="stSidebar"] .stCheckbox label { color: #b0c8d8 !important; font-size:0.82rem !important; }

    /* ── Hero ── */
    .hero {
        background: linear-gradient(135deg, #1e3a5f 0%, #0d2137 60%, #0a3355 100%);
        border-radius: 16px; padding: 2rem 2.5rem 1.5rem;
        margin-bottom: 1.5rem; box-shadow: 0 8px 32px rgba(0,0,0,0.18);
        position: relative; overflow: hidden;
    }
    .hero::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 200px; height: 200px; background: rgba(96,200,240,0.08);
        border-radius: 50%;
    }
    .hero-title {
        font-family: 'Inter', sans-serif; font-size: 2rem; font-weight: 700;
        color: #ffffff; letter-spacing: -0.5px; margin-bottom: 0.3rem;
    }
    .hero-title span { color: #60c8f0; }
    .hero-sub {
        font-family: 'JetBrains Mono', monospace; color: #6a90a8;
        font-size: 0.75rem; margin-bottom: 1rem;
    }
    .hero-brand {
        display: inline-block;
        background: rgba(96,200,240,0.12); border: 1px solid rgba(96,200,240,0.25);
        border-radius: 20px; padding: 0.35rem 1rem;
        font-family: 'Inter', sans-serif; font-size: 0.78rem; color: #a0d8f0;
        font-weight: 500;
    }
    .hero-brand b { color: #60c8f0; }

    /* ── Cards ── */
    .metric-card {
        background: #ffffff; border: 1px solid #dce8f0;
        border-radius: 12px; padding: 1.2rem 1rem; text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }
    .metric-val   { font-family: 'Inter', sans-serif; font-size: 2rem; font-weight: 700; color: #1e3a5f; }
    .metric-label { font-family: 'Inter', sans-serif; color: #7a9ab0; font-size: 0.68rem; text-transform: uppercase; letter-spacing: 0.08em; margin-top: 0.2rem; }
    .metric-val.red   { color: #e05555; }
    .metric-val.amber { color: #d48b00; }
    .metric-val.green { color: #2a9d5c; }

    /* ── Alert boxes ── */
    .tip  { background:#e8f8f2; border-left:4px solid #2a9d5c; border-radius:8px; padding:0.7rem 1rem; font-size:0.82rem; color:#1a6640; margin-bottom:0.8rem; }
    .info { background:#e8f2fc; border-left:4px solid #3b82c4; border-radius:8px; padding:0.7rem 1rem; font-size:0.82rem; color:#1a3f7a; margin-bottom:0.8rem; }
    .warn { background:#fff8e8; border-left:4px solid #d48b00; border-radius:8px; padding:0.7rem 1rem; font-size:0.82rem; color:#7a4f00; margin-bottom:0.5rem; }
    .err  { background:#fdf0f0; border-left:4px solid #e05555; border-radius:8px; padding:0.7rem 1rem; font-size:0.82rem; color:#8b1a1a; margin-bottom:0.5rem; }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab-list"] { background: #dce8f4; border-radius: 10px; padding: 4px; }
    .stTabs [data-baseweb="tab"] { border-radius: 8px; font-family: 'Inter', sans-serif; font-weight: 600; color: #5a7a90; font-size: 0.85rem; }
    .stTabs [aria-selected="true"] { background: #ffffff !important; color: #1e3a5f !important; box-shadow: 0 2px 6px rgba(0,0,0,0.1); }

    /* ── Buttons ── */
    div[data-testid="stDownloadButton"] button {
        background: linear-gradient(135deg, #1e3a5f, #2a5280) !important;
        color: white !important; border: none !important; font-weight: 600 !important;
        border-radius: 8px !important; font-family: 'Inter', sans-serif !important;
        box-shadow: 0 3px 10px rgba(30,58,95,0.3) !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #1e7a4f, #2a9d5c) !important;
        font-weight: 700 !important; border-radius: 10px !important;
        box-shadow: 0 4px 14px rgba(42,157,92,0.35) !important;
    }

    /* ── Data editor ── */
    .stDataFrame, [data-testid="stDataEditor"] {
        border-radius: 10px !important; border: 1px solid #dce8f0 !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05) !important;
    }

    /* ── Footer ── */
    .footer {
        text-align: center; padding: 1.5rem;
        font-family: 'Inter', sans-serif; font-size: 0.75rem; color: #8aabb8;
        border-top: 1px solid #dce8f0; margin-top: 2rem;
    }
    .footer b { color: #1e3a5f; }
</style>
""", unsafe_allow_html=True)

# ── Hero Banner ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div class="hero-title">⚡ Data<span>Scan</span> → Excel</div>
  <div class="hero-sub">// AI Vision · Smart SKU Matching · Ditto Fill · Error Detection</div>
  <div class="hero-brand">
    Powered by &nbsp;<b>Ashu Bhatt</b>&nbsp;·&nbsp;Accounts &amp; Finance Department&nbsp;·&nbsp;<b>Yash Gallery Private Limited</b>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Session state helper ─────────────────────────────────────────────────────
def get_sku_set():
    return st.session_state.get("master_skus", set())

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Groq API Key")
    st.markdown("""
    <div class="info">
    🆓 <b>FREE — No Credit Card</b><br><br>
    1. <a href="https://console.groq.com" target="_blank" style="color:#60c8f0"><b>console.groq.com</b></a><br>
    2. Sign up → API Keys → Create<br>
    3. Paste below ↓
    </div>
    """, unsafe_allow_html=True)
    default_key = ""
    try: default_key = st.secrets.get("GROQ_API_KEY", "")
    except: pass
    api_key = st.text_input("Groq API Key", value=default_key, type="password", placeholder="gsk_...")

    st.markdown("---")
    st.markdown("### 📋 Columns")
    col_input = st.text_area("Columns (comma-separated)", value="SKU,QTY,BIN", height=80)
    columns = [c.strip() for c in col_input.split(",") if c.strip()]

    st.markdown("---")
    st.markdown("### 📂 Master SKU File")
    sku_file = st.file_uploader("Upload Test.xlsx", type=["xlsx"])
    if sku_file:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(sku_file)
            ws = wb.active
            skus = [str(r[0]).strip() for r in ws.iter_rows(min_row=4, values_only=True) if r[0] and str(r[0]).strip()]
            st.session_state["master_skus"] = set(skus)
            st.success(f"✅ {len(skus):,} SKUs loaded!")
        except Exception as e:
            st.error(f"Error: {e}")

    st.markdown("---")
    st.markdown("### 🎯 Options")
    ditto_fill   = st.checkbox('Smart ditto fill (" → copy above)', value=True)
    validate_sku = st.checkbox("Validate & fix SKUs against master list", value=True)
    enhance_img  = st.checkbox("Enhance image before extraction", value=True)

    st.markdown("---")
    st.markdown("""<div style='font-size:0.7rem;color:#5a7a90;'>
    <b style='color:#60c8f0'>SKU Formats Supported:</b><br>
    <code style='color:#a0d8f0'>1001YKBEIGE-XL</code><br>
    <code style='color:#a0d8f0'>AK-103BLUE-XXL</code><br>
    <code style='color:#a0d8f0'>1003KDMUSTARD-11-12</code><br>
    <code style='color:#a0d8f0'>7001YKBLS-L-XL</code><br>
    <code style='color:#a0d8f0'>4001DRSRED-S</code><br>
    <code style='color:#a0d8f0'>6003SKDGREEN-XL</code><br>
    <code style='color:#a0d8f0'>1613-XL → auto resolved</code>
    </div>""", unsafe_allow_html=True)

# ── Image helpers ─────────────────────────────────────────────────────────────
def enhance_image(pil_img):
    img = pil_img.convert("RGB")
    w, h = img.size
    MAX_DIM = 1400
    if w < 1200:
        scale = min(1200 / w, MAX_DIM / max(w, h, 1))
        img = img.resize((int(w*scale), int(h*scale)), Image.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(1.6)
    img = ImageEnhance.Sharpness(img).enhance(1.5)
    img = ImageEnhance.Brightness(img).enhance(1.05)
    return img

def image_to_base64(pil_img):
    buf = io.BytesIO()
    pil_img.save(buf, format="JPEG", quality=95)
    return base64.b64encode(buf.getvalue()).decode()

# ── AI Prompt ─────────────────────────────────────────────────────────────────
def build_prompt(columns):
    col_list = ", ".join(columns)
    return f"""You are a careful OCR reader for handwritten Indian clothing warehouse ledger sheets.
Read EXACTLY what is written. Never invent or guess values.

COLUMNS TO EXTRACT: {col_list}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
THE SKU COLUMN HAS TWO WRITTEN PARTS — READ BOTH
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
In this sheet the SKU column is split into two parts written side by side:

  PART A (left):  The BASE — a number + code like:  1592 YK  |  1539 YK  |  AK-141
  PART B (right): The SIZE — written after = or " like:  = XL  |  " XL  |  = XXL  |  = M

The FULL SKU = BASE + "-" + SIZE
Examples:
  PART A="1592 YK"   PART B="= XL"   →  SKU = 1592YK-XL
  PART A="1539 YK"   PART B="= XL"   →  SKU = 1539YK-XL
  PART A="1884 YK 83" PART B="= XL"  →  SKU = 1884YK83-XL
  PART A="AK-141"    PART B="= XL"   →  SKU = AK-141-XL
  PART A="1517 YK"   PART B="= M"    →  SKU = 1517YK-M

DITTO RULE FOR PART A (BASE) ONLY:
  If PART A shows:  "  or  \  or  //  or  〃  or is blank
  → Copy the BASE from the row above (just the NUMBER+CODE part, NOT the size)
  → Then attach the SIZE from PART B of the current row

  Example:
    Row 1: PART A="1592 YK"   PART B="= XL"   →  1592YK-XL
    Row 2: PART A="  "  (blank/ditto)  PART B="= XXL"  →  1592YK-XXL
    Row 3: PART A="1539 YK"   PART B="= XL"   →  1539YK-XL

  IMPORTANT: The ditto " only copies the BASE (number+code). 
  The SIZE always comes from PART B of the CURRENT row.
  NEVER copy the full previous SKU including size when size differs.

HOW TO BUILD THE BASE:
  - Read the digits: 1592, 1539, 1162, 1884, 1338, 5006, 6019, 1255, 1739, 1426
  - Read the code immediately after: YK, KD, SKD, DRS, MW, YKBLS, DPT
  - If a number appears between code and color like "YK83" or "YK148" include it
  - Remove all spaces: "1592 YK" → "1592YK"
  - Do NOT add any color unless a color word is clearly written in that row
  - Color words: BLACK BLUE RED WHITE GREEN MAROON MUSTARD BEIGE GREY NAVY etc.
  - If no color word is written → SKU has no color part

HOW TO READ THE SIZE (PART B):
  = XL   →  XL
  = XXL  →  XXL
  " XL   →  XL  (the " here means same size as above, or just read the size written)
  = M    →  M
  = S    →  S
  = L    →  L
  = 3XL  →  3XL

SIZE OPTIONS: XS S M L XL XXL 3XL 4XL 5XL 6XL 7XL 8XL
  Kids: 7-8 9-10 11-12 13-14 | Combo: S-M L-XL XXL-3XL F

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
BIN COLUMN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
BIN is a shelf location. It has multiple parts:  T5 - A7 - D4
Read ALL parts and join with dashes → T5-A7-D4

The " symbol or ~ or blank in BIN → copy FULL BIN from row above.
BIN is written once for a group of rows. All rows below inherit it until a new BIN appears.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
QTY COLUMN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Read the number carefully. 0≠O, 1≠l.
"9+9" or "2+3" → output as written.
" or blank in QTY → copy from row above.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
WORKED EXAMPLE — matches the real sheet exactly
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PDF line 1:  "1592 YK"  "= XL"   QTY=1    BIN="T5 - A7 - D4"
PDF line 2:  "1539 YK"  "= XL"   QTY=9+9  BIN=" "
PDF line 3:  "1162 YK"  "= XL"   QTY=1    BIN=" "
PDF line 4:  "1884 YK83" "= XL"  QTY=1    BIN=" "
PDF line 5:  "1338 YK"  "= XL"   QTY=1    BIN=" "
PDF line 6:  (blank)    "= XL"   QTY=1    BIN=" "   ← blank BASE = ditto 1338YK

Correct output:
[
  {{"SKU": "1592YK-XL",   "QTY": "1",   "BIN": "T5-A7-D4"}},
  {{"SKU": "1539YK-XL",   "QTY": "9+9", "BIN": "T5-A7-D4"}},
  {{"SKU": "1162YK-XL",   "QTY": "1",   "BIN": "T5-A7-D4"}},
  {{"SKU": "1884YK83-XL", "QTY": "1",   "BIN": "T5-A7-D4"}},
  {{"SKU": "1338YK-XL",   "QTY": "1",   "BIN": "T5-A7-D4"}},
  {{"SKU": "1338YK-XL",   "QTY": "1",   "BIN": "T5-A7-D4"}}
]

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. ONE output row per ONE handwritten line — no more, no less
2. Ditto on BASE → copy base from above + use current row SIZE
3. Never copy the size from the previous row unless current row size is also ditto
4. Never invent a color if not written
5. Full BIN always — never just the last part like "D4"
6. UPPERCASE for all SKU and BIN values
7. No spaces inside SKU

Return ONLY a raw JSON array. No markdown. No explanation.
"""


# ── Groq API call ─────────────────────────────────────────────────────────────
def extract_with_groq(api_key, pil_img, columns, _retry=0):
    """
    Call Groq API with automatic retry on rate-limit (429).
    Retries up to 5 times with exponential back-off.
    """
    MAX_RETRIES = 5
    b64 = image_to_base64(pil_img)
    payload = {
        "model": "meta-llama/llama-4-scout-17b-16e-instruct",
        "messages": [{"role": "user", "content": [
            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
            {"type": "text", "text": build_prompt(columns)}
        ]}],
        "max_tokens": 8192, "temperature": 0
    }
    resp = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json=payload, timeout=120
    )

    # ── Rate limit: auto-retry with back-off ──────────────────────────────────
    if resp.status_code == 429:
        if _retry >= MAX_RETRIES:
            raise ValueError(f"Groq rate limit hit — exceeded {MAX_RETRIES} retries. Try uploading fewer pages at once.")
        # Parse retry-after from response if available
        try:
            err_body = resp.json()
            msg = err_body.get("error", {}).get("message", "")
            # Extract "Please try again in X.Xs" if present
            m = re.search(r"try again in ([\d.]+)s", msg)
            wait = float(m.group(1)) if m else 2 ** (_retry + 1)
        except Exception:
            wait = 2 ** (_retry + 1)

        wait = max(wait + 1.5, 3.0)   # always add buffer
        st.warning(f"⏳ Groq rate limit hit — waiting {wait:.1f}s before retry {_retry+1}/{MAX_RETRIES}…")
        time.sleep(wait)
        return extract_with_groq(api_key, pil_img, columns, _retry=_retry + 1)

    if resp.status_code != 200:
        raise ValueError(f"Groq API error {resp.status_code}: {resp.text[:300]}")

    raw = resp.json()["choices"][0]["message"]["content"].strip()
    raw = re.sub(r"```(?:json)?|```", "", raw).strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        if m: data = json.loads(m.group())
        else: raise ValueError(f"Could not parse JSON:\n{raw[:400]}")
    return data if isinstance(data, list) else []

# ── SKU Lookup Maps ───────────────────────────────────────────────────────────
SIZE_SUFFIX_RE = re.compile(
    r'-(XS|S|M|L|XL|XXL|[3-8]XL|F|F-S/XXL|F-3XL/5XL|\d+-\d+|S-M|L-XL|XXL-3XL|4XL-5XL|XS-S|M-L|S/M|L/XL|XS/S|M/L)$',
    re.IGNORECASE
)

def build_lookup_maps(master_skus):
    """Build 5 lookup structures for fast SKU resolution."""
    base_map     = {}   # BASE_SKU (no size) → [full SKUs]
    num_size_map = {}   # (num_str, SIZE) → [full SKUs]
    num_map      = {}   # num_str → [full SKUs]
    prefix3_map  = {}   # first 3 digits → [full SKUs]  ← NEW
    prefix4_map  = {}   # first 4 digits → [full SKUs]  ← NEW

    for s in master_skus:
        base = SIZE_SUFFIX_RE.sub("", s).upper()
        base_map.setdefault(base, []).append(s)

        m_num  = re.match(r"^(\d+)", s)
        m_size = SIZE_SUFFIX_RE.search(s)
        if m_num:
            num_str = m_num.group(1)
            num_map.setdefault(num_str, []).append(s)
            # 3-digit prefix (e.g. "143" covers 1430, 1431…)
            if len(num_str) >= 3:
                p3 = num_str[:3]
                prefix3_map.setdefault(p3, []).append(s)
            # 4-digit prefix
            if len(num_str) >= 4:
                p4 = num_str[:4]
                prefix4_map.setdefault(p4, []).append(s)
            if m_size:
                key = (num_str, m_size.group(1).upper())
                num_size_map.setdefault(key, []).append(s)

    return base_map, num_size_map, num_map, prefix3_map, prefix4_map


def _clean(s):
    return re.sub(r"\s+", "", s).upper()


def _suggest_similar(base_in, base_map, n=3):
    """Return up to n fuzzy suggestions from master list for error rows."""
    hits = get_close_matches(base_in, list(base_map.keys()), n=n, cutoff=0.55)
    suggestions = []
    for h in hits:
        candidates = sorted(base_map[h])
        suggestions.append(candidates[0])
    return suggestions


def validate_and_fix_sku(sku, master_skus, base_map, num_size_map, num_map,
                          prefix3_map=None, prefix4_map=None):
    """
    Returns (corrected_sku, status, note, expanded_list, suggestions)
    status: 'ok' | 'fixed' | 'expanded' | 'error'
    suggestions: list of similar SKUs to show when status=='error'
    """
    prefix3_map  = prefix3_map  or {}
    prefix4_map  = prefix4_map  or {}
    raw = str(sku).strip()
    if not raw or raw == "nan":
        return raw, "ok", "", [], []

    # 1. Exact match
    if raw in master_skus:
        return raw, "ok", "", [], []

    u = _clean(raw)

    # 2. Uppercase / whitespace fix
    if u in master_skus:
        return u, "fixed", "case/space fix", [], []

    def _pick(candidates):
        if len(candidates) == 1: return candidates[0], "fixed", "resolved", [], []
        if len(candidates) > 1:  return candidates[0], "expanded", f"multi({len(candidates)})", candidates, []
        return None, "error", "", [], []

    # 3. NUM-SIZE with dash/space: "1613-XL", "1536 XL"
    m = re.match(r"^(\d+)[\s-]+(.+)$", u)
    if m:
        num, size = m.group(1), m.group(2).strip("-").strip()
        c = num_size_map.get((num, size))
        if c:
            r, st_, note, ex, sg = _pick(c); 
            if r: return r, st_, note, ex, sg
        for tok in [size.split("-")[0], size.split("-")[-1]]:
            c2 = num_size_map.get((num, tok))
            if c2:
                r, st_, note, ex, sg = _pick(c2)
                if r: return r, st_, note, ex, sg

    # 4. NUMSIZE no dash: "1536XL"
    m2 = re.match(r"^(\d+)(XS|XXL|XL|[3-8]XL|S(?!KD)|M(?!W)|L(?!AVE))$", u)
    if m2:
        num, size = m2.group(1), m2.group(2)
        c = num_size_map.get((num, size))
        if c:
            r, st_, note, ex, sg = _pick(c)
            if r: return r, st_, note, ex, sg

    # 5. Base exact match (no size) → expand all sizes
    base_in = SIZE_SUFFIX_RE.sub("", u)
    if base_in in base_map:
        full = sorted(base_map[base_in])
        m_sz = SIZE_SUFFIX_RE.search(u)
        if m_sz:
            sz   = m_sz.group(1).upper()
            same = [s for s in full if s.upper().endswith("-" + sz)]
            if same: return same[0], "fixed", "base+size", [], []
        return full[0], "expanded", f"expanded({len(full)} sizes)", full, []

    # 6. ── Prefix match — ONLY when color is genuinely missing from OCR output ──
    # Do NOT auto-fix if the OCR gave a color — trust what was read.
    # Only trigger when the SKU has no color part (just NUMBER+CODE+SIZE).
    m_num = re.match(r"^(\d+)", u)
    # Detect if color is absent: SKU matches pattern like 1592YK-XL or 1592YK (no color word)
    has_no_color = bool(re.match(r"^(\d+)(YK|KD|SKD|DRS|MW|YKBLS|DPT)([-]?(XS|S|M|L|XL|XXL|[3-8]XL|F|\d+-\d+|S-M|L-XL)?)?$", u, re.IGNORECASE))
    if m_num and has_no_color:
        num_str = m_num.group(1)
        m_sz = SIZE_SUFFIX_RE.search(u)
        sz = m_sz.group(1).upper() if m_sz else None

        # Try 4-digit prefix — only if EXACTLY ONE candidate matches size
        p4 = num_str[:4] if len(num_str) >= 4 else None
        if p4 and p4 in prefix4_map:
            candidates = sorted(prefix4_map[p4])
            if sz:
                same = [s for s in candidates if s.upper().endswith("-" + sz)]
                if len(same) == 1:
                    return same[0], "fixed", f"4-digit prefix match ({p4})", [], []
                elif len(same) > 1:
                    return same[0], "expanded", f"4-digit prefix ({p4}, {len(same)} options)", same, []
            else:
                if len(candidates) == 1:
                    return candidates[0], "fixed", f"4-digit prefix match ({p4})", [], []
                return candidates[0], "expanded", f"4-digit prefix ({p4}, {len(candidates)} SKUs)", candidates, []

        # Try 3-digit prefix — only if EXACTLY ONE candidate matches size
        p3 = num_str[:3] if len(num_str) >= 3 else None
        if p3 and p3 in prefix3_map:
            candidates = sorted(prefix3_map[p3])
            if sz:
                same = [s for s in candidates if s.upper().endswith("-" + sz)]
                if len(same) == 1:
                    return same[0], "fixed", f"3-digit prefix match ({p3})", [], []
                elif len(same) > 1:
                    return same[0], "expanded", f"3-digit prefix ({p3}, {len(same)} options)", same, []
            else:
                if len(candidates) == 1:
                    return candidates[0], "fixed", f"3-digit prefix match ({p3})", [], []
                return candidates[0], "expanded", f"3-digit prefix ({p3}, {len(candidates)} SKUs)", candidates, []

    # 7. Fuzzy base match
    bm = get_close_matches(base_in, list(base_map.keys()), n=1, cutoff=0.80)
    if bm:
        full = sorted(base_map[bm[0]])
        m_sz = SIZE_SUFFIX_RE.search(u)
        if m_sz:
            sz   = m_sz.group(1).upper()
            same = [s for s in full if s.upper().endswith("-" + sz)]
            if same: return same[0], "fixed", "fuzzy+size", [], []
        return full[0], "expanded", f"fuzzy expanded({len(full)})", full, []

    # 8. Fuzzy full-SKU last resort
    hits = get_close_matches(u, list(master_skus), n=1, cutoff=0.82)
    if hits: return hits[0], "fixed", "fuzzy match", [], []

    # ❌ Not found — generate suggestions
    suggestions = _suggest_similar(base_in, base_map, n=3)
    return raw, "error", "NOT FOUND IN MASTER LIST", [], suggestions


def apply_ditto(df):
    """
    Correct ditto fill logic matching real handwritten sheet structure:

    The SKU column is built from TWO parts:
      BASE  = the number+code (e.g. 1592YK)
      SIZE  = the size suffix (e.g. XL, XXL, M)

    When SKU cell has a ditto mark (" \ // blank etc.):
      → Copy only the BASE from the row above
      → Keep the SIZE from the CURRENT row (don't copy size from above)
      → Recombine: BASE + "-" + SIZE

    BIN is forward-filled (written once, covers all rows below).
    QTY ditto copies the value from the row above.
    Truly blank rows (no SKU, no QTY) are dropped.
    """
    import re as _re

    SIZE_RE = _re.compile(
        r'-(XS|S|M|L|XL|XXL|[3-8]XL|F|\d+-\d+|S-M|L-XL|XXL-3XL|4XL-5XL|XS-S|M-L)$',
        _re.IGNORECASE
    )

    def is_ditto(val):
        v = str(val).strip()
        if not v or v.lower() in ("nan", "none"):
            return True
        if _re.match(r'^["\\/,`\'~=]{1,3}$', v):
            return True
        if v in ('"', "''", "//", ",,", "\\", "\u3003", "11", "~", "="):
            return True
        return False

    def is_truly_blank(val):
        v = str(val).strip()
        return v == "" or v.lower() in ("nan", "none")

    def get_base(sku):
        """Strip size suffix to get just the base: 1592YK-XL → 1592YK"""
        return SIZE_RE.sub("", str(sku).strip())

    def get_size(sku):
        """Extract size suffix: 1592YK-XL → XL"""
        m = SIZE_RE.search(str(sku).strip())
        return m.group(1) if m else ""

    df = df.copy()

    # Mark truly blank rows BEFORE any fill (all data cols empty/nan)
    data_cols = [c for c in df.columns if c.upper() not in ("PDF PAGE", "SOURCE FILE")]
    truly_empty_idx = df.index[
        df[data_cols].apply(lambda row: all(is_truly_blank(v) for v in row), axis=1)
    ].tolist()

    # ── Pass 1: Ditto fill ────────────────────────────────────────────────────
    for col in df.columns:
        for i in range(1, len(df)):
            val = str(df.at[i, col]).strip()
            if not is_ditto(val):
                continue
            above = df.at[i-1, col]
            if is_truly_blank(above):
                continue

            if col.upper() == "SKU":
                # Ditto on SKU = copy BASE from above, keep SIZE from current row
                above_base = get_base(str(above))
                above_size = get_size(str(above))
                raw_current = str(df.at[i, col]).strip()
                cur_size = get_size(raw_current) if not is_ditto(raw_current) else ""
                # Use current row's size if it has one, else fall back to above size
                size = cur_size if cur_size else above_size
                if size:
                    df.at[i, col] = above_base + "-" + size
                else:
                    df.at[i, col] = above_base
            else:
                df.at[i, col] = above

    # ── Pass 2: BIN forward-fill ──────────────────────────────────────────────
    for col in df.columns:
        if col.upper() == "BIN":
            last_bin = None
            for i in range(len(df)):
                val = df.at[i, col]
                if not is_ditto(val) and not is_truly_blank(val):
                    last_bin = val
                elif last_bin is not None:
                    df.at[i, col] = last_bin

    # ── Pass 3: Drop truly blank rows ─────────────────────────────────────────
    df = df.drop(index=truly_empty_idx, errors="ignore").reset_index(drop=True)

    return df


# ── Excel export ──────────────────────────────────────────────────────────────
def to_excel(df, row_status):
    """
    row_status: dict of row_index -> 'ok' | 'fixed' | 'expanded' | 'error'
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extracted Data")
        ws = writer.sheets["Extracted Data"]

        # ── Colour palette ──
        CLR_HEADER_BG   = "1E3A5F"
        CLR_HEADER_FG   = "FFFFFF"
        CLR_ROW_EVEN    = "F5F9FC"
        CLR_ROW_ODD     = "FFFFFF"
        CLR_FIXED_BG    = "FFF8E1"   # soft amber
        CLR_FIXED_FG    = "7A5000"
        CLR_EXPANDED_BG = "E8F5FF"   # light blue
        CLR_EXPANDED_FG = "0D4A80"
        CLR_ERROR_BG    = "FFF0F0"   # light red
        CLR_ERROR_FG    = "C0392B"
        CLR_BORDER      = "D0DDE8"

        border = Border(
            bottom=Side(style="thin", color=CLR_BORDER),
            right=Side(style="thin", color=CLR_BORDER)
        )

        # ── Header row ──
        for ci, cell in enumerate(ws[1], 1):
            cell.fill  = PatternFill("solid", fgColor=CLR_HEADER_BG)
            cell.font  = Font(bold=True, color=CLR_HEADER_FG, name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            max_len = max((len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row+1)), default=10)
            ws.column_dimensions[cell.column_letter].width = min(max_len + 4, 50)

        ws.row_dimensions[1].height = 24

        # ── Find SKU and SKU Status column indices ──
        sku_col_idx    = None
        status_col_idx = None
        sugg_col_idx   = None
        page_col_idx   = None
        for ci, cell in enumerate(ws[1], 1):
            v = str(cell.value).strip().upper()
            if v == "SKU":         sku_col_idx    = ci
            if v == "SKU STATUS":  status_col_idx = ci
            if v == "SUGGESTIONS": sugg_col_idx   = ci
            if v == "PDF PAGE":    page_col_idx   = ci

        # ── Data rows ──
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            row_idx  = ri - 2
            status   = row_status.get(row_idx, "ok")

            # Base row alternating colour
            bg = CLR_ROW_EVEN if ri % 2 == 0 else CLR_ROW_ODD
            base_fill = PatternFill("solid", fgColor=bg)

            for cell in row:
                cell.fill      = base_fill
                cell.font      = Font(name="Calibri", color="2C3E50", size=10)
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

            # Override SKU cell colour by status
            if sku_col_idx:
                sku_cell = ws.cell(ri, sku_col_idx)
                if status == "fixed":
                    sku_cell.fill = PatternFill("solid", fgColor=CLR_FIXED_BG)
                    sku_cell.font = Font(name="Calibri", color=CLR_FIXED_FG, size=10, bold=True)
                elif status == "expanded":
                    sku_cell.fill = PatternFill("solid", fgColor=CLR_EXPANDED_BG)
                    sku_cell.font = Font(name="Calibri", color=CLR_EXPANDED_FG, size=10, bold=True)
                elif status == "error":
                    sku_cell.fill = PatternFill("solid", fgColor=CLR_ERROR_BG)
                    sku_cell.font = Font(name="Calibri", color=CLR_ERROR_FG, size=10, bold=True)

            # Colour the SKU Status cell too
            if status_col_idx:
                sc = ws.cell(ri, status_col_idx)
                if status == "fixed":
                    sc.fill = PatternFill("solid", fgColor=CLR_FIXED_BG)
                    sc.font = Font(name="Calibri", color=CLR_FIXED_FG, size=10)
                elif status == "expanded":
                    sc.fill = PatternFill("solid", fgColor=CLR_EXPANDED_BG)
                    sc.font = Font(name="Calibri", color=CLR_EXPANDED_FG, size=10)
                elif status == "error":
                    sc.fill = PatternFill("solid", fgColor=CLR_ERROR_BG)
                    sc.font = Font(name="Calibri", color=CLR_ERROR_FG, size=10, bold=True)

            # Colour Suggestions cell (orange bg when it has content)
            if sugg_col_idx:
                sgc = ws.cell(ri, sugg_col_idx)
                if sgc.value and str(sgc.value).strip():
                    sgc.fill = PatternFill("solid", fgColor="FFF3CD")
                    sgc.font = Font(name="Calibri", color="7A5000", size=9, italic=True)

            # Style PDF Page cell — light grey, centered, monospace-like
            if page_col_idx:
                pgc = ws.cell(ri, page_col_idx)
                pgc.fill = PatternFill("solid", fgColor="F0F4F8")
                pgc.font = Font(name="Courier New", color="3A5A7A", size=9)
                pgc.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        # ── Legend sheet ──
        lg = writer.book.create_sheet("Legend")
        legend_data = [
            ("Colour / Column", "Meaning"),
            ("White / Alternating", "SKU matched exactly in master list"),
            ("Amber background", "SKU was auto-corrected / resolved (check the SKU Status column for details)"),
            ("Blue background", "SKU was a partial — expanded to all matching sizes"),
            ("Red background", "❌ SKU NOT FOUND — please review manually"),
            ("Suggestions column", "Similar SKUs from master list — pick the correct one for red rows"),
            ("PDF Page column", "Which page of the PDF this row came from — use to verify against original"),
        ]
        for r, (a, b) in enumerate(legend_data, 1):
            lg.cell(r, 1, a); lg.cell(r, 2, b)
            if r == 1:
                for c in [lg.cell(r,1), lg.cell(r,2)]:
                    c.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
                    c.font = Font(bold=True, color="FFFFFF", name="Calibri")
            elif r == 3:
                lg.cell(r,1).fill = PatternFill("solid", fgColor=CLR_FIXED_BG)
                lg.cell(r,1).font = Font(color=CLR_FIXED_FG, name="Calibri")
            elif r == 4:
                lg.cell(r,1).fill = PatternFill("solid", fgColor=CLR_EXPANDED_BG)
                lg.cell(r,1).font = Font(color=CLR_EXPANDED_FG, name="Calibri")
            elif r == 5:
                lg.cell(r,1).fill = PatternFill("solid", fgColor=CLR_ERROR_BG)
                lg.cell(r,1).font = Font(color=CLR_ERROR_FG, name="Calibri", bold=True)
        lg.column_dimensions["A"].width = 28
        lg.column_dimensions["B"].width = 50

    buf.seek(0)
    return buf.read()


# ── UI ────────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📤 Upload & Extract", "📊 Data & Export"])

with tab1:
    master_skus = get_sku_set()
    if master_skus:
        st.markdown(f'<div class="tip">✅ Master SKU list loaded — <b>{len(master_skus):,} SKUs</b> ready for validation.</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="warn">⚠️ No master SKU file loaded. Upload <b>Test.xlsx</b> in the sidebar to enable SKU validation.</div>', unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Drop your handwritten sheet images or PDFs here",
        type=["jpg","jpeg","png","webp","pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed"
    )

    if uploaded_files:
        col_go, _ = st.columns([1, 3])
        with col_go:
            go = st.button("⚡ Extract All", type="primary", use_container_width=True)

        if go:
            if not api_key:
                st.error("⚠️ Enter your Groq API Key in the sidebar.")
            elif not columns:
                st.error("⚠️ Add at least one column.")
            else:
                all_rows  = []
                progress  = st.progress(0, text="Starting...")

                for idx, uf in enumerate(uploaded_files):
                    progress.progress(idx / len(uploaded_files), text=f"Processing {uf.name}…")
                    try:
                        file_bytes = uf.read()
                        is_pdf = uf.name.lower().endswith(".pdf")

                        if is_pdf:
                            # Convert each PDF page to an image
                            if not PDF_SUPPORT:
                                st.error(f"❌ PDF support not available. Install pdf2image + poppler.")
                                continue
                            # Write to temp file to avoid holding full bytes in RAM
                            tmp_path = None
                            file_rows = []
                            try:
                                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                                    tmp.write(file_bytes)
                                    tmp_path = tmp.name
                                del file_bytes  # free original bytes immediately
                                gc.collect()

                                # Count pages first at low DPI for speed
                                probe = convert_from_path(tmp_path, dpi=30, first_page=1, last_page=1)
                                total_pages = len(convert_from_path(tmp_path, dpi=30))
                                del probe
                                gc.collect()

                                st.info(f"📄 `{uf.name}` — {total_pages} page(s) found")
                                pg_status = st.empty()
                                rows_ticker = st.empty()

                                # Free tier = 30,000 TPM. Each page ≈ 2,000 tokens.
                                # 4s gap → ~15 pages/min → safely under limit
                                THROTTLE_SECONDS = 4.0

                                for pg_num in range(1, total_pages + 1):
                                    pg_status.info(f"⏳ Processing page {pg_num} / {total_pages}  |  {len(file_rows)} rows saved so far…")
                                    # Load ONE page at a time at safe DPI
                                    page_imgs = convert_from_path(
                                        tmp_path, dpi=130,
                                        first_page=pg_num, last_page=pg_num,
                                        fmt="jpeg"
                                    )
                                    if not page_imgs:
                                        continue
                                    page_img = page_imgs[0]
                                    img = page_img.convert("RGB")
                                    page_img.close()
                                    del page_imgs
                                    if enhance_img:
                                        img = enhance_image(img)

                                    try:
                                        pg_rows = extract_with_groq(api_key, img, columns)
                                    except Exception as page_err:
                                        pg_status.error(f"❌ Page {pg_num} failed: {page_err} — skipping, continuing with next page.")
                                        img.close(); del img; gc.collect()
                                        # Save partial progress so user doesn't lose already-extracted rows
                                        if file_rows:
                                            st.session_state["partial_rows"] = file_rows.copy()
                                        time.sleep(THROTTLE_SECONDS)
                                        continue

                                    img.close()
                                    del img
                                    gc.collect()

                                    for r in pg_rows:
                                        r["__source"] = uf.name
                                        r["__page"] = str(pg_num)
                                    file_rows.extend(pg_rows)

                                    # Save partial progress after every page
                                    st.session_state["partial_rows"] = file_rows.copy()

                                    rows_ticker.success(f"✅ Page {pg_num}/{total_pages} → {len(pg_rows)} rows  |  Total: {len(file_rows)}")

                                    # Throttle between pages to avoid 429
                                    if pg_num < total_pages:
                                        time.sleep(THROTTLE_SECONDS)

                            finally:
                                if tmp_path and os.path.exists(tmp_path):
                                    os.unlink(tmp_path)

                            all_rows.extend(file_rows)
                            st.success(f"✅ `{uf.name}` — total {len(file_rows)} rows from {total_pages} pages")
                        else:
                            import io as _io
                            img = Image.open(_io.BytesIO(file_bytes))
                            del file_bytes
                            gc.collect()
                            if enhance_img: img = enhance_image(img)
                            rows = extract_with_groq(api_key, img, columns)
                            img.close()
                            del img
                            gc.collect()
                            for r in rows: r["__source"] = uf.name; r["__page"] = "1"
                            all_rows.extend(rows)
                            st.success(f"✅ `{uf.name}` → {len(rows)} rows extracted")

                    except Exception as e:
                        st.error(f"❌ `{uf.name}`: {e}")

                progress.progress(1.0, text="Done!")

                if all_rows:
                    df = pd.DataFrame(all_rows)
                    for c in columns:
                        if c not in df.columns: df[c] = ""
                    source_col = df.pop("__source") if "__source" in df.columns else None
                    page_col   = df.pop("__page")   if "__page"   in df.columns else None
                    df = df[columns]

                    if ditto_fill:
                        df = apply_ditto(df)

                    row_status = {}   # row_index -> 'ok'|'fixed'|'expanded'|'error'

                    # SKU Validation + Smart Resolution
                    if validate_sku and master_skus and "SKU" in df.columns:
                        base_map, num_size_map, num_map, prefix3_map, prefix4_map = build_lookup_maps(master_skus)

                        fixed_count    = 0
                        expanded_count = 0
                        error_count    = 0
                        new_rows       = []

                        for i, row in df.iterrows():
                            sku = str(row.get("SKU", "")).strip()
                            if not sku:
                                new_rows.append({**row.to_dict(), "SKU Status": "—", "Suggestions": ""})
                                row_status[len(new_rows)-1] = "ok"
                                continue

                            corrected, status, note, expanded, suggestions = validate_and_fix_sku(
                                sku, master_skus, base_map, num_size_map, num_map,
                                prefix3_map, prefix4_map
                            )

                            sugg_str = " | ".join(suggestions) if suggestions else ""

                            if expanded and len(expanded) > 1:
                                for full_sku in expanded:
                                    new_rows.append({**row.to_dict(), "SKU": full_sku,
                                                     "SKU Status": f"✅ Expanded from: {sku}", "Suggestions": ""})
                                    row_status[len(new_rows)-1] = "expanded"
                                expanded_count += len(expanded)
                                fixed_count    += 1

                            elif status == "fixed":
                                new_rows.append({**row.to_dict(), "SKU": corrected,
                                                 "SKU Status": f"✅ Fixed ({note}): {sku} → {corrected}", "Suggestions": ""})
                                row_status[len(new_rows)-1] = "fixed"
                                fixed_count += 1

                            elif status == "error":
                                new_rows.append({**row.to_dict(), "SKU": sku,
                                                 "SKU Status": f"❌ NOT FOUND: {sku}",
                                                 "Suggestions": sugg_str})
                                row_status[len(new_rows)-1] = "error"
                                error_count += 1

                            else:
                                new_rows.append({**row.to_dict(), "SKU Status": "✅ OK", "Suggestions": ""})
                                row_status[len(new_rows)-1] = "ok"

                        df = pd.DataFrame(new_rows)

                        # Re-order columns: original columns + SKU Status + Suggestions + Page + Source
                        all_cols = [c for c in columns if c in df.columns]
                        for extra in ["SKU Status", "Suggestions", "PDF Page", "Source File"]:
                            if extra in df.columns and extra not in all_cols:
                                all_cols.append(extra)
                        if page_col is not None and "PDF Page" not in df.columns:
                            df["PDF Page"] = page_col.values if len(page_col) == len(df) else ""
                        if source_col is not None:
                            if "Source File" not in df.columns:
                                df["Source File"] = ""
                            if len(source_col) == len(df):
                                df["Source File"] = source_col.values
                        df = df[[c for c in all_cols if c in df.columns]]

                        # Summary alerts
                        if fixed_count or expanded_count:
                            msg = []
                            if fixed_count:    msg.append(f"**{fixed_count} SKUs** auto-resolved")
                            if expanded_count: msg.append(f"**{expanded_count} rows** generated from partials")
                            st.markdown(f'<div class="warn">🔧 {" · ".join(msg)} — highlighted in Excel.</div>', unsafe_allow_html=True)
                        if error_count:
                            st.markdown(f'<div class="err">❌ <b>{error_count} SKUs</b> could NOT be matched — check the <b>SKU Status</b> column (red rows in Excel).</div>', unsafe_allow_html=True)

                    else:
                        # No validation — still add status column
                        df["SKU Status"] = "—"
                        df["Suggestions"] = ""
                        if page_col is not None:
                            df["PDF Page"] = page_col.values if len(page_col)==len(df) else ""
                        if source_col is not None:
                            df["Source File"] = source_col.values if len(source_col)==len(df) else ""

                    st.session_state["df"]         = df
                    st.session_state["row_status"] = row_status
                    st.session_state["ready"]      = True
                    st.balloons()
                    st.info(f"✨ **{len(df)} rows** extracted — go to **📊 Data & Export** tab.")
                else:
                    st.warning("No rows extracted. Check image quality or try again.")
                    # Offer to recover partial rows if any were saved mid-run
                    if st.session_state.get("partial_rows"):
                        n = len(st.session_state["partial_rows"])
                        st.info(f"💾 **{n} rows** were saved before the error. Click below to recover them.")
                        if st.button("♻️ Recover Partial Results"):
                            all_rows = st.session_state["partial_rows"]
                            st.session_state.pop("partial_rows", None)
                            # Re-run the rest of the pipeline with recovered rows
                            st.rerun()

with tab2:
    if st.session_state.get("ready") and "df" in st.session_state:
        df         = st.session_state["df"]
        row_status = st.session_state.get("row_status", {})

        # ── Metrics ──
        total    = len(df)
        ok_cnt   = sum(1 for v in row_status.values() if v == "ok")
        fixed    = sum(1 for v in row_status.values() if v in ("fixed","expanded"))
        errors   = sum(1 for v in row_status.values() if v == "error")
        src_cnt  = df["Source File"].nunique() if "Source File" in df.columns else 1

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1: st.markdown(f'<div class="metric-card"><div class="metric-val">{total}</div><div class="metric-label">Total Rows</div></div>', unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="metric-card"><div class="metric-val green">{ok_cnt}</div><div class="metric-label">Exact Match</div></div>', unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="metric-card"><div class="metric-val amber">{fixed}</div><div class="metric-label">Auto Fixed</div></div>', unsafe_allow_html=True)
        with c4: st.markdown(f'<div class="metric-card"><div class="metric-val red">{errors}</div><div class="metric-label">❌ Errors</div></div>', unsafe_allow_html=True)
        with c5: st.markdown(f'<div class="metric-card"><div class="metric-val">{src_cnt}</div><div class="metric-label">Source Files</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Error highlight notice ──
        if errors > 0:
            st.markdown(f'<div class="err">❌ <b>{errors} rows</b> have unmatched SKUs — see <b>SKU Status</b> column below. These appear <b>red</b> in the downloaded Excel.</div>', unsafe_allow_html=True)

        # ── Filter toggle ──
        show_errors_only = st.checkbox("🔍 Show only unmatched SKU rows", value=False)
        display_df = df.copy()
        if show_errors_only and "SKU Status" in display_df.columns:
            display_df = display_df[display_df["SKU Status"].str.startswith("❌", na=False)]

        st.markdown("#### 📋 Extracted Data *(click any cell to edit)*")
        edited_df = st.data_editor(display_df, use_container_width=True, num_rows="dynamic", hide_index=False)

        if show_errors_only:
            st.caption("⚠️ Showing filtered view — download uses full dataset.")
        else:
            st.session_state["df"] = edited_df

        st.markdown("---")

        # ── Colour legend in app ──
        st.markdown("""
        <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:1rem;">
          <span style="background:#FFF8E1;border:1px solid #D4A800;border-radius:6px;padding:4px 12px;font-size:0.78rem;color:#7A5000;font-weight:600;">🟡 Amber = Auto-Fixed</span>
          <span style="background:#E8F5FF;border:1px solid #3b82c4;border-radius:6px;padding:4px 12px;font-size:0.78rem;color:#0D4A80;font-weight:600;">🔵 Blue = Expanded from partial</span>
          <span style="background:#FFF0F0;border:1px solid #E05555;border-radius:6px;padding:4px 12px;font-size:0.78rem;color:#C0392B;font-weight:600;">🔴 Red = NOT FOUND — review needed</span>
          <span style="background:#F5F9FC;border:1px solid #D0DDE8;border-radius:6px;padding:4px 12px;font-size:0.78rem;color:#2C3E50;font-weight:600;">⚪ White = Exact match ✅</span>
        </div>
        """, unsafe_allow_html=True)

        d1, d2, d3 = st.columns([1,1,2])
        with d1:
            export_df = st.session_state["df"]
            st.download_button(
                "⬇️ Download Excel (.xlsx)",
                data=to_excel(export_df, row_status),
                file_name="extracted_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with d2:
            st.download_button(
                "⬇️ Download CSV",
                data=st.session_state["df"].to_csv(index=False).encode(),
                file_name="extracted_data.csv",
                mime="text/csv",
                use_container_width=True
            )
        with d3:
            if st.button("🗑️ Clear & Reset", use_container_width=True):
                for k in ["df","ready","row_status"]: st.session_state.pop(k, None)
                st.rerun()
    else:
        st.markdown("""
        <div style="text-align:center;padding:5rem 2rem;color:#8aabb8;">
          <div style="font-size:3.5rem">📊</div>
          <div style="font-size:1.1rem;font-weight:600;color:#1e3a5f;margin:1rem 0 0.5rem">No data yet</div>
          <div style="font-size:0.85rem">Upload images and click <b style="color:#2a9d5c">⚡ Extract All</b></div>
        </div>
        """, unsafe_allow_html=True)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
  Powered by &nbsp;<b>Ashu Bhatt</b>&nbsp;·&nbsp;Accounts &amp; Finance Department&nbsp;·&nbsp;<b>Yash Gallery Private Limited</b><br>
  <span style="font-size:0.68rem;color:#a0b8c8;">Built with Streamlit · AI by Groq · SKU Validation Engine v3</span>
</div>
""", unsafe_allow_html=True)
